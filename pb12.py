from openpyxl import  Workbook
wb = Workbook()

#inside the wb --->which sheet we need to work with
#grab the active worksheet
ws = wb.active
#active sheet--> sheet that is open or mention the sheet name

#data can be assigned directly to cells
ws['A1'] = 42

#rows can also be appended
ws.append([1, 2, 3])

#python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

#save the file
wb.save("sample1.xlsx")